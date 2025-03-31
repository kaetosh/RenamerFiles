import os
import openpyxl
import time
import sys

class EmptyFiles(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)
# Путь к папке с файлами Excel
folder_path = os.getcwd()

type_file = input('Введите расширение файлов, которых необходимо переименовать.\nНапример, для excel файлов введите xlsx, для word - docx и т.д.\n')
print(f'Убедитесь, что в папке существуют файлы .{type_file}, которые требуется переименовать.')
input('если все ОК, нажмите <Enter>, иначе - закройте программу\n')

# Создаем файл Excel для переименования
try:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Rename Files'

    # Формируем первую колонку со старыми названиями файлов
    files = [f for f in os.listdir(folder_path) if (f.endswith(f'.{type_file}') and f != 'rename_files.xlsx')]
    if not files:
        raise EmptyFiles("Файлы с указанным расширением не найдены!!!")
    for i, file in enumerate(files):
        sheet.cell(row=i+1, column=1, value=file)

    # Формируем вторую колонку для новых названий файлов
    for i in range(len(files)):
        sheet.cell(row=i+1, column=2, value='')

    # Сохраняем файл Excel
    wb.save(f'{folder_path}/rename_files.xlsx')
    time.sleep(1)
    print('Cоздан excel файл <rename_files.xlsx> с именами файлов в столбце A.')
except EmptyFiles as e:
    print(f"Произошла ошибка: {e.message}")
    input('Закройте программу.\n')
    sys.exit()
except Exception as e:
    print(f'Ошибка при создании или сохранении Excel файла: {e}')
    input('Закройте программу.\n')
    sys.exit()

print('Заполните столбец B с новыми названиями файлов. Сохраните файл.')
input('Как будете готовы, нажмите <Enter>\n')

# Второй этап: переименование файлов
try:
    wb = openpyxl.load_workbook(f'{folder_path}/rename_files.xlsx', data_only=True)
    sheet = wb.active

    for row in sheet.iter_rows(values_only=True):
        old_name, new_name = row
        if new_name:
            try:
                os.rename(f'{folder_path}/{old_name}', f'{folder_path}/{new_name}')
                print(f'Файл {old_name} переименован в {new_name}')
            except FileNotFoundError:
                print(f'Ошибка: Файл {old_name} не найден.')
                input('Закройте программу.\n')
                sys.exit()
            except FileExistsError:
                print(f'Ошибка: Файл {new_name} уже существует.')
                input('Закройте программу.\n')
                sys.exit()
            except Exception as e:
                print(f'Ошибка при переименовании файла {old_name}: {e}')
                input('Закройте программу.\n')
                sys.exit()
    time.sleep(1)
    input('Закройте программу.\n')
    sys.exit()
except Exception as e:
    print(f'Ошибка при загрузке Excel файла: {e}')
    input('Закройте программу.\n')
    sys.exit()

